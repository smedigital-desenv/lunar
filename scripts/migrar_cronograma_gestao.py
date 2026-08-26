#!/usr/bin/env python3
# =====================================================================
# migrar_cronograma_gestao.py — converte o "Cronograma das Gerências"
# da Subsecretaria de Gestão Administrativa, Financeira e Tecnológica
# (planilha .xlsx) em SQL para o schema gestao.
#
# Formato da planilha: uma aba por gerência; dentro de cada aba, doze
# blocos mensais (JANEIRO…DEZEMBRO), e em cada bloco as colunas
# AÇÃO | RESPONSÁVEL | STATUS | DATA DE CONCLUSÃO | OBSERVAÇÕES.
#
# Desenho acertado com o usuário (2026-08-26):
#   · a mesma AÇÃO se repete mês a mês (a Zeladoria tem 15 ações nos 12
#     meses = 125 linhas). Uma demanda por linha encheria o painel de
#     repetições, então **cada ação distinta vira UMA demanda** e cada
#     mês em que ela aparece vira **UMA TAREFA**, com o prazo daquele
#     mês. É o que o cronograma de "Meus processos" foi feito para
#     mostrar;
#   · a DEMANDA fica na unidade do titular indicado (`EMAIL_TITULAR`),
#     que é solicitante e responsável — como na migração da Pedagógica;
#   · a TAREFA vai para o titular da GERÊNCIA da aba;
#   · autor/solicitante da carga = usuário de serviço "Dados importados"
#     — ninguém real assina o que não escreveu;
#   · objeto_queixa (NOT NULL, sem coluna de origem) repete o título.
#
# ⚠️ O RESPONSÁVEL da planilha NÃO é usado para atribuir a tarefa. A
# coluna traz primeiro nome ou rótulo de setor ("Educ-Fin",
# "T.I/Desenvolvimento"), que não casa com o nome completo do cadastro;
# casar por primeiro nome atribuiria trabalho à pessoa errada, e errar
# atribuindo é invisível. O texto original fica registrado na descrição
# da tarefa, e o gerente redistribui.
#
# ⚠️ O STATUS de origem também fica registrado por extenso. O banco só
# conhece aberta/em_andamento/concluída; a planilha usa dezesseis
# grafias ("executando", "Respondendo apontamentos da PGM", "SUSPENSO").
# Reduzir sem preservar apagaria o que a gerência escreveu.
#
# A tarefa é criada no formato "subtarefa direto na demanda" (sql/023):
# parent_id nulo e SEM mover responsavel_atual_id — que é o estado que a
# tela produz. fn_encaminhar moveria o responsável e tiraria a demanda
# da Subsecretaria, o oposto do combinado.
#
# O SQL gerado NÃO é versionado (dados reais, repositório público).
#
# Uso:
#   python3 scripts/migrar_cronograma_gestao.py <planilha.xlsx> \\
#          <email-do-titular@educacao.pmrp.sp.gov.br> [pasta-de-saida]
# =====================================================================

import calendar
import os
import sys
import unicodedata
from datetime import date, datetime

import openpyxl

ANO = 2026

# Colunas do bloco mensal (índice 0). A coluna A é uma calha vazia.
C_ACAO, C_RESP, C_STATUS, C_DATA, C_OBS = 1, 2, 3, 4, 5

MESES = ['JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO', 'JULHO',
         'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']
# ⚠️ Comparação de mês é SEMPRE pela forma normalizada: chave('MARÇO') é
# 'MARCO'. Testar contra MESES cru faz o bloco de março não ser
# reconhecido — ele vira uma "ação" e as linhas dele entram em fevereiro.
MESES_CHAVE = [' '.join(unicodedata.normalize('NFD', m).encode('ascii', 'ignore')
                        .decode().upper().split()) for m in MESES]

ABA_IGNORADA = 'ACOMPANHAMENTO'      # só cabeçalho, sem dado

# Quantas demandas cabem num arquivo. Cada uma custa ~3,5 kB de SQL, então
# 40 dá ~140 kB — o que o SQL Editor do Supabase aguenta colado de uma vez.
MAX_DEMANDAS_POR_ARQUIVO = 40

# ⚠️ O e-mail do titular NÃO fica neste arquivo: ele é dado pessoal, e
# este repositório é público. Vem na linha de comando e só aparece no SQL
# gerado, que o .gitignore barra. Foi um punhado de endereços
# @educacao.pmrp.sp.gov.br versionados que tirou os .sql do Git em
# 2026-08-26 (sql/LEIA-ME.md).
#
# A unidade também não é embutida: sai da lotação do titular no cadastro,
# então mudança de organograma não exige mexer aqui.
EMAIL_TITULAR = None      # preenchido por principal(), a partir do argumento

# Usuário de serviço da migração (criado na recarga de Licitações).
EMAIL_AUTOR = 'dados.importados@educacao.pmrp.sp.gov.br'

# Status da planilha -> situacao do banco. A chave é comparada sem acento
# e sem caixa. O que não estiver aqui aborta a conversão: status novo tem
# de ser classificado por gente, não cair num padrão silencioso.
SITUACAO = {
    'concluido':                        'concluida',
    'em andamento':                     'em_andamento',
    'em desenvolvimento':               'em_andamento',
    'executando':                       'em_andamento',
    'em processo de compra':            'em_andamento',
    'em processo de compra / licitacao': 'em_andamento',
    'respondendo apontamentos da pgm':  'em_andamento',
    'reuniao para alinhamento':         'em_andamento',
    'aguardando':                       'aberta',
    'aguardando inicio':                'aberta',
    'nao iniciado':                     'aberta',
    'suspenso':                         'aberta',
    'em definicao':                     'aberta',
    '':                                 'aberta',   # célula vazia
}

# Grafias da planilha que não batem com o organograma. A busca já ignora
# acento e caixa; isto é para diferença de PALAVRAS mesmo. Deixar em
# branco não é opção silenciosa: nome que não resolver aborta a carga
# inteira, no pré-voo, antes de gravar qualquer coisa.
TEXTO_CRIACAO = ('Demanda migrada do Cronograma das Gerências {ano} ({gerencia}). '
                 'Cada mês da planilha virou uma tarefa. A data de criação é o '
                 '1º dia do mês da primeira ocorrência.').replace('{ano}', str(ANO))

MAPA_UNIDADES = {
    # 'NOME NA PLANILHA': 'Nome exato no organograma',
}


# ------------------------------------------------------------- utilidades
def sem_acento(txt):
    return ''.join(c for c in unicodedata.normalize('NFD', str(txt))
                   if unicodedata.category(c) != 'Mn')


def chave(txt):
    """Forma comparável: sem acento, sem caixa, espaços colapsados."""
    return ' '.join(sem_acento(txt or '').upper().split())


def norm(expr):
    """Expressão SQL que normaliza um nome para comparação — a mesma de
    scripts/migrar_subped.py. Aspas são enfeite de grafia, não identidade
    da unidade (o organograma escreve Centro de Formação "Paulo Freire"
    entre aspas e as planilhas não)."""
    sem_aspas = f"""translate({expr}, '"''´`', '')"""
    return f"btrim(regexp_replace(upper(unaccent({sem_aspas})), '\\s+', ' ', 'g'))"


def sql_txt(v):
    """Literal SQL. None vira NULL; aspas simples são dobradas."""
    if v is None:
        return 'null'
    return "'" + str(v).replace("'", "''") + "'"


def sql_data(d):
    return f"'{d.isoformat()}'" if d else 'null'


def limpar(v):
    """Texto da célula sem espaços sobrando. O traço solitário que a
    planilha usa como 'nada aqui' vira ausência de verdade."""
    if v is None:
        return None
    t = ' '.join(str(v).split())
    if t in ('', '-', '--'):
        return None
    return t


def mapear(nome):
    alvo = chave(nome)
    for de, para in MAPA_UNIDADES.items():
        if chave(de) == alvo:
            return para
    return nome


def como_data(v, onde, avisos):
    """Data da planilha. O que não for data reconhecível não vira palpite:
    devolve None e sai no relatório — a planilha tem pelo menos um
    '30/02/2026', que não existe no calendário."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    t = limpar(v)
    if not t:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(t, fmt).date()
        except ValueError:
            pass
    avisos.append(f'{onde}: data inválida na planilha ({t!r}) — usado o fim do mês.')
    return None


def fundir(alvo, nova):
    """Junta duas linhas iguais do mesmo mês sem perder nada do que elas
    dizem de diferente. A situação mais adiantada vence — 'em andamento'
    ao lado de 'aguardando' significa que aquilo andou."""
    ordem = {'aberta': 0, 'em_andamento': 1, 'concluida': 2}
    if ordem[nova['situacao']] > ordem[alvo['situacao']]:
        alvo['situacao'] = nova['situacao']
    for campo in ('status_origem', 'responsavel_origem', 'obs'):
        a, b = alvo[campo], nova[campo]
        if b and b != a:
            alvo[campo] = f'{a} / {b}' if a else b
    if nova['data'] and (not alvo['data'] or nova['data'] > alvo['data']):
        alvo['data'] = nova['data']
        alvo['prazo'] = nova['data']
    alvo['onde'] = f"{alvo['onde']} (+ {nova['onde'].split()[-1]})"


def fim_do_mes(mes_idx):
    return date(ANO, mes_idx, calendar.monthrange(ANO, mes_idx)[1])


# ------------------------------------------------------------- leitura
def ler(caminho):
    """Devolve (acoes, avisos, vazias).

    `acoes` é uma lista de dicionários — uma por (gerência, ação) — com
    as ocorrências mensais dentro, na ordem do calendário."""
    wb = openpyxl.load_workbook(caminho, data_only=True)
    avisos, vazias = [], []
    acoes = {}                       # (gerencia, chave(acao)) -> dict

    for ws in wb.worksheets:
        if ws.title == ABA_IGNORADA:
            continue
        gerencia = limpar(ws.cell(1, 2).value)
        if not gerencia:
            avisos.append(f'Aba {ws.title!r} sem nome de gerência na célula B1 — ignorada.')
            continue

        mes_idx = None
        achou = 0
        for linha, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if linha == 1:
                continue                                   # título da aba
            celulas = [limpar(v) for v in row]
            acao = celulas[C_ACAO] if len(celulas) > C_ACAO else None
            if not acao:
                continue
            if chave(acao) in MESES_CHAVE:
                mes_idx = MESES_CHAVE.index(chave(acao)) + 1
                continue
            if chave(acao) == 'ACAO':
                continue                                   # cabeçalho do bloco
            if mes_idx is None:
                avisos.append(f'{ws.title} linha {linha}: ação fora de qualquer bloco '
                              f'mensal — ignorada ({acao[:50]!r}).')
                continue

            onde = f'{ws.title} linha {linha}'
            status = celulas[C_STATUS] if len(celulas) > C_STATUS else None
            situacao = SITUACAO.get(chave(status).lower())
            if situacao is None:
                sys.exit(f'{onde}: status desconhecido {status!r}. '
                         f'Classifique-o em SITUACAO (scripts/migrar_cronograma_gestao.py). '
                         f'Conhecidos: {sorted(k for k in SITUACAO if k)}')

            # Valor CRU: openpyxl devolve datetime, e limpar() o transformaria
            # em texto '2026-01-27 00:00:00', que nenhum formato casa.
            data = como_data(row[C_DATA] if len(row) > C_DATA else None, onde, avisos)
            obs = celulas[C_OBS] if len(celulas) > C_OBS else None

            k = (gerencia, chave(acao))
            it = acoes.setdefault(k, {
                'gerencia': gerencia,
                'gerencia_banco': mapear(gerencia),
                'titulo': acao,
                'ocorrencias': [],
            })
            nova = {
                'onde': onde,
                'mes': mes_idx,
                'mes_nome': MESES[mes_idx - 1].capitalize(),
                'situacao': situacao,
                'status_origem': status,
                'responsavel_origem': celulas[C_RESP] if len(celulas) > C_RESP else None,
                'data': data,
                'prazo': data or fim_do_mes(mes_idx),
                'obs': obs,
            }
            # A planilha repete linhas dentro do MESMO mês (o bloco de
            # janeiro da Prestação de Contas foi colado duas vezes). Duas
            # tarefas com o mesmo título na mesma demanda seriam uma só
            # depois da guarda de idempotência — a segunda sumiria em
            # silêncio, e o total do relatório mentiria. Aqui elas se
            # FUNDEM, e a fusão é anunciada.
            anterior = next((o for o in it['ocorrencias'] if o['mes'] == mes_idx), None)
            if anterior is None:
                it['ocorrencias'].append(nova)
                achou += 1
            else:
                avisos.append(f'{onde}: linha repetida de {anterior["onde"]} '
                              f'(mesma ação, mesmo mês) — fundidas numa tarefa só.')
                fundir(anterior, nova)

        if achou == 0:
            vazias.append(gerencia)

    for it in acoes.values():
        it['ocorrencias'].sort(key=lambda o: o['mes'])
    return list(acoes.values()), avisos, vazias


# ------------------------------------------------------------- consolidação
def consolidar(it):
    """Situação, prazo e conclusão da DEMANDA a partir dos meses dela."""
    ocs = it['ocorrencias']
    situacoes = {o['situacao'] for o in ocs}

    if situacoes == {'concluida'}:
        datas = [o['data'] for o in ocs if o['data']]
        textos = [f"{o['mes_nome']}: {o['obs']}" for o in ocs if o['obs']]
        it['situacao'] = 'concluida'
        it['prazo'] = None
        it['data_conclusao'] = max(datas) if datas else max(o['prazo'] for o in ocs)
        # chk_demanda_conclusao exige texto. Quando a planilha não traz
        # observação (50 das 54 linhas concluídas), o texto diz de onde a
        # conclusão veio, em vez de inventar um relato que ninguém escreveu.
        it['conclusao'] = ('\n'.join(textos) if textos else
                           'Registrada como concluída no cronograma das gerências. '
                           'A planilha de origem não traz texto de conclusão.')
    else:
        it['situacao'] = 'em_andamento' if 'em_andamento' in situacoes else 'aberta'
        # Prazo da demanda = o ÚLTIMO mês ainda em aberto, não o primeiro.
        # A ação recorrente é um compromisso do ano: com o primeiro, a
        # demanda de roçada nasceria vencida em janeiro e as 195 abertas
        # apareceriam todas atrasadas no dia da carga. O mês a mês fica
        # nas tarefas, que é onde ele é cobrado.
        abertas = [o['prazo'] for o in ocs if o['situacao'] != 'concluida']
        it['prazo'] = max(abertas) if abertas else None
        it['data_conclusao'] = None
        it['conclusao'] = None

    meses = ', '.join(o['mes_nome'] for o in it['ocorrencias'])
    partes = [f"Ação do cronograma {ANO} da {it['gerencia']}.",
              f"Meses previstos na planilha: {meses}."]
    resp = sorted({o['responsavel_origem'] for o in ocs if o['responsavel_origem']})
    if resp:
        partes.append('Responsável na planilha de origem: ' + '; '.join(resp) + '.')
    it['descricao'] = '\n'.join(partes)


def descricao_tarefa(oc):
    partes = [f"{oc['mes_nome']}/{ANO} do cronograma das gerências."]
    if oc['responsavel_origem']:
        partes.append(f"Responsável na planilha de origem: {oc['responsavel_origem']}.")
    if oc['status_origem']:
        partes.append(f"Situação na planilha de origem: {oc['status_origem']}.")
    if oc['obs']:
        partes.append(f"Observações: {oc['obs']}")
    return '\n'.join(partes)


def conclusao_tarefa(oc):
    if oc['situacao'] != 'concluida':
        return None
    return oc['obs'] or ('Registrada como concluída no cronograma das gerências. '
                         'A planilha de origem não traz texto de conclusão.')


# ------------------------------------------------------------- geração
def cabecalho(planilha, n_dem, n_tar, escopo):
    return f"""-- Cronograma das Gerências — Subsecretaria de Gestão Administrativa,
-- Financeira e Tecnológica. {escopo}
-- Gerado por scripts/migrar_cronograma_gestao.py a partir de
-- {planilha}
-- {n_dem} demanda(s) e {n_tar} tarefa(s) — uma tarefa por mês de cada ação.
-- NÃO versionar: contém dados reais.
--
-- Idempotente por (titulo, unidade_responsavel_id) na demanda e por
-- (demanda_id, titulo) na tarefa: rodar de novo não duplica. Nada é
-- apagado (regra 1 — não existe DELETE).
--
-- ⚠️ O SQL Editor do Supabase envolve CADA comando numa transação: se um
-- bloco falhar, o que ele gravou é desfeito por inteiro. Os blocos são
-- independentes entre si — o que já passou fica.

set search_path = gestao, extensions, public;
"""


def bloco_gerencia(gerencia, acoes):
    """Um bloco DO por gerência. A unidade e o titular dela são resolvidos
    UMA vez, não uma por ação: com 181 ações em Orçamento e Finanças, a
    repetição sozinha triplicava o tamanho do arquivo."""
    n_tar = sum(len(it['ocorrencias']) for it in acoes)
    nome_banco = acoes[0]['gerencia_banco']
    out = [f"""
-- =====================================================================
-- {gerencia} — {len(acoes)} demanda(s), {n_tar} tarefa(s)
-- =====================================================================
do $$
declare
  v_autor    uuid;
  v_titular  uuid;
  v_unidade  uuid;
  v_ger      uuid;
  v_resp     uuid;
  v_demanda  uuid;
  v_tarefa   uuid;
  v_novas    int := 0;
  v_pulou    int := 0;
begin
  select id into v_autor from gestao.usuarios where email = {sql_txt(EMAIL_AUTOR)};
  if v_autor is null then
    raise exception 'Usuário de serviço da migração não existe: %', {sql_txt(EMAIL_AUTOR)};
  end if;

  select id, unidade_id into v_titular, v_unidade
    from gestao.usuarios where email = {sql_txt(EMAIL_TITULAR)} and ativo;
  if v_titular is null then
    raise exception 'Titular não encontrado (ou inativo) no cadastro: %', {sql_txt(EMAIL_TITULAR)};
  end if;
  if v_unidade is null then
    raise exception 'Titular % não tem unidade de lotação — a demanda ficaria sem dono.',
      {sql_txt(EMAIL_TITULAR)};
  end if;

  -- A gerência que recebe as tarefas deste bloco.
  select id into v_ger from gestao.unidades_organizacionais
   where {norm('nome')} = {norm(sql_txt(nome_banco))} and ativo;
  if v_ger is null then
    raise exception 'Gerência não encontrada no organograma: % — confira o nome com o SQL de conferência e mapeie a grafia em MAPA_UNIDADES (scripts/migrar_cronograma_gestao.py). Nada deste bloco foi gravado.',
      {sql_txt(nome_banco)};
  end if;

  -- Titular da gerência: quem está lotado nela com o maior nível.
  -- Inativo conta como ausente: atribuir a quem saiu ressuscitaria trabalho.
  select u.id into v_resp
    from gestao.usuarios u join gestao.perfis p on p.codigo = u.perfil
   where u.unidade_id = v_ger and u.ativo
   order by p.nivel desc, u.criado_em limit 1;
  if v_resp is null then
    -- Sem titular a tarefa não some: fica com o subsecretário, que
    -- redistribui. Sumir em silêncio é que ela não pode.
    raise notice 'Sem titular ativo em % — as tarefas ficaram com o subsecretário.',
      {sql_txt(nome_banco)};
    v_resp := v_titular;
  end if;"""]

    for it in acoes:
        primeira = it['ocorrencias'][0]
        criado = date(ANO, primeira['mes'], 1)
        rotulo = it['titulo'][:70].replace(chr(10), ' ')
        out.append(f"""
  -- {rotulo}
  select id into v_demanda from gestao.demandas
   where titulo = {sql_txt(it['titulo'])} and unidade_responsavel_id = v_unidade;

  if v_demanda is null then
    insert into gestao.demandas(
      titulo, objeto_queixa, descricao, situacao, prioridade, prazo,
      conclusao, data_conclusao, sigilo,
      solicitante_id, responsavel_atual_id, unidade_responsavel_id,
      criado_por, criado_em)
    values (
      {sql_txt(it['titulo'])}, {sql_txt(it['titulo'])}, {sql_txt(it['descricao'])},
      {sql_txt(it['situacao'])}, 'normal', {sql_data(it['prazo'])}::date,
      {sql_txt(it['conclusao'])}, {sql_data(it['data_conclusao'])}::timestamptz, 'normal',
      v_titular, v_titular, v_unidade, v_autor, {sql_data(criado)}::timestamptz)
    returning id into v_demanda;

    -- Mesmos participantes que fn_criar_demanda registraria.
    perform gestao.fn_garantir_participante(v_demanda, v_titular, 'solicitante', v_autor);
    perform gestao.fn_garantir_participante(v_demanda, v_resp, 'participante', v_autor);

    insert into gestao.movimentacoes(
      demanda_id, autor_id, tipo, texto, situacao_nova, destinatario_id,
      prioridade, criado_em)
    values (v_demanda, v_autor, 'criacao', {sql_txt(TEXTO_CRIACAO.format(gerencia=acoes[0]['gerencia']))},
      {sql_txt(it['situacao'])}, v_titular, 'normal', {sql_data(criado)}::timestamptz);
    v_novas := v_novas + 1;
  else
    v_pulou := v_pulou + 1;
  end if;""")

        for oc in it['ocorrencias']:
            titulo_t = f"{it['titulo']} — {oc['mes_nome']}/{ANO}"
            criado_t = date(ANO, oc['mes'], 1)
            concluida = oc['situacao'] == 'concluida'
            out.append(f"""
  -- {oc['onde']} · {oc['mes_nome']}
  if not exists (select 1 from gestao.tarefas
                  where demanda_id = v_demanda and titulo = {sql_txt(titulo_t)}) then
    insert into gestao.tarefas(
      demanda_id, parent_id, titulo, descricao, responsavel_id,
      unidade_responsavel_id, situacao, prioridade, prazo,
      conclusao, data_conclusao, criado_por, criado_em)
    values (v_demanda, null, {sql_txt(titulo_t)}, {sql_txt(descricao_tarefa(oc))},
      v_resp, v_ger, {sql_txt(oc['situacao'])}, 'normal',
      {sql_data(None if concluida else oc['prazo'])}::date,
      {sql_txt(conclusao_tarefa(oc))},
      {sql_data(oc['data'] or oc['prazo'] if concluida else None)}::timestamptz,
      v_autor, {sql_data(criado_t)}::timestamptz)
    returning id into v_tarefa;

    insert into gestao.movimentacoes(
      demanda_id, tarefa_id, autor_id, tipo, texto, situacao_nova,
      destinatario_id, prioridade, criado_em)
    values (v_demanda, v_tarefa, v_autor, 'subtarefa', {sql_txt(titulo_t)},
      {sql_txt(oc['situacao'])}, v_resp, 'normal', {sql_data(criado_t)}::timestamptz);
  end if;""")

    out.append(f"""
  raise notice '{gerencia.replace("'", "''")}: % demanda(s) nova(s), % já existiam.',
    v_novas, v_pulou;
end $$;
""")
    return '\n'.join(out)


def gerar_conferencia(acoes):
    """SQL só de leitura, para rodar ANTES: mostra o que o banco encontra.
    Sem isso a carga só falha no pré-voo, e cada nome ruim custa uma
    rodada."""
    nomes = sorted({it['gerencia_banco'] for it in acoes})
    vals = ', '.join(f'({sql_txt(n)})' for n in nomes)
    return f"""-- Conferência de nomes — SÓ LEITURA. Rode ANTES da carga.
-- Toda linha com achou = false precisa de correção (na planilha, no
-- organograma ou em MAPA_UNIDADES) antes de migrar.
set search_path = gestao, extensions, public;

-- 1) As gerências existem no organograma, e têm titular para receber tarefa?
select 'gerencia' as tipo, v.nome,
       (u.id is not null) as achou,
       (select count(*) from gestao.usuarios x
         where x.unidade_id = u.id and x.ativo) as pessoas_ativas
  from (values {vals}) as v(nome)
  left join gestao.unidades_organizacionais u
    on {norm('u.nome')} = {norm('v.nome')} and u.ativo
order by achou, nome;

-- 2) Titular do cronograma e autor da carga.
select 'titular' as tipo, {sql_txt(EMAIL_TITULAR)} as email,
       exists (select 1 from gestao.usuarios
                where email = {sql_txt(EMAIL_TITULAR)} and ativo) as achou,
       (select un.nome from gestao.usuarios us
          join gestao.unidades_organizacionais un on un.id = us.unidade_id
         where us.email = {sql_txt(EMAIL_TITULAR)}) as unidade
union all
select 'autor', {sql_txt(EMAIL_AUTOR)},
       exists (select 1 from gestao.usuarios where email = {sql_txt(EMAIL_AUTOR)}), null;

-- 3) Já existe demanda com algum destes títulos? (a carga PULA as que
--    existirem — nada é duplicado, mas nada é atualizado também)
select count(*) as titulos_ja_existentes
  from gestao.demandas d
 where d.titulo in ({', '.join(sql_txt(it['titulo']) for it in acoes)});
"""


def sufixo(gerencia):
    """Nome de arquivo a partir do nome da gerência."""
    base = chave(gerencia).lower().replace('gerencia de ', '').replace('gerencia ', '')
    return ''.join(c if c.isalnum() else '-' for c in base).strip('-').replace('--', '-')


def principal():
    global EMAIL_TITULAR
    if len(sys.argv) < 3:
        sys.exit('Uso: python3 scripts/migrar_cronograma_gestao.py '
                 '<planilha.xlsx> <email-do-titular> [pasta]\n'
                 'O e-mail não tem padrão: ele é dado pessoal e não mora neste arquivo.')
    planilha = sys.argv[1]
    EMAIL_TITULAR = sys.argv[2].strip()
    if not EMAIL_TITULAR.endswith('@educacao.pmrp.sp.gov.br'):
        sys.exit(f'E-mail fora do domínio da rede: {EMAIL_TITULAR!r} — '
                 'o banco recusaria o cadastro (trg_bloquear_dominio).')
    pasta = sys.argv[3] if len(sys.argv) > 3 else 'db'
    os.makedirs(pasta, exist_ok=True)

    acoes, avisos, vazias = ler(planilha)
    if not acoes:
        sys.exit('Nenhuma ação encontrada na planilha — nada a migrar.')
    for it in acoes:
        consolidar(it)

    por_ger = {}
    for it in acoes:
        por_ger.setdefault(it['gerencia'], []).append(it)

    conf = os.path.join(pasta, '00-conferencia.sql')
    with open(conf, 'w', encoding='utf-8') as f:
        f.write(gerar_conferencia(acoes))
    print(f'conferência de nomes (rode PRIMEIRO) -> {conf}')
    print()

    # Um arquivo por gerência: o SQL Editor não engasga, e uma falha numa
    # gerência não desfaz o que as outras já gravaram. Gerência grande é
    # partida em pedaços — colar 600 kB num editor de navegador trava a
    # aba, e o arquivo que ninguém consegue rodar não migrou nada.
    ordem = 0
    for ger, itens in sorted(por_ger.items()):
        partes = [itens[k:k + MAX_DEMANDAS_POR_ARQUIVO]
                  for k in range(0, len(itens), MAX_DEMANDAS_POR_ARQUIVO)] or [[]]
        for n, parte in enumerate(partes, start=1):
            ordem += 1
            n_tar = sum(len(x['ocorrencias']) for x in parte)
            marca = f'-parte{n}' if len(partes) > 1 else ''
            escopo = ger + (f' (parte {n} de {len(partes)})' if len(partes) > 1 else '')
            caminho = os.path.join(pasta, f'{ordem:02d}-{sufixo(ger)}{marca}.sql')
            with open(caminho, 'w', encoding='utf-8') as f:
                f.write(cabecalho(planilha, len(parte), n_tar, escopo))
                f.write(bloco_gerencia(escopo, parte))
            kb = os.path.getsize(caminho) // 1024
            print(f'  {escopo}: {len(parte):3d} demandas, {n_tar:3d} tarefas '
                  f'-> {caminho} ({kb} kB)')

    print()
    for s_ in sorted({it['situacao'] for it in acoes}):
        print(f'  demandas {s_}: {sum(1 for it in acoes if it["situacao"] == s_)}')
    for s_ in sorted({o['situacao'] for it in acoes for o in it['ocorrencias']}):
        print(f'  tarefas  {s_}: '
              f'{sum(1 for it in acoes for o in it["ocorrencias"] if o["situacao"] == s_)}')
    if vazias:
        print()
        print('  Gerências SEM nenhuma ação na planilha (nada foi criado para elas):')
        for g in vazias:
            print(f'    - {g}')
    if avisos:
        print()
        print('  Avisos:')
        for a in avisos:
            print(f'    - {a}')


if __name__ == '__main__':
    principal()
