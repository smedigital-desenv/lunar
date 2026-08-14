# =====================================================================
# migrar_licitacoes.py — converte a planilha "Acompanhamento Licitações
# e Contratos" (3 abas de compras) em SQL para o módulo de Licitações.
#
# Uso:  python3 scripts/migrar_licitacoes.py <planilha.xlsx> [saida.sql]
# Requer: pip install openpyxl
#
# O SQL gerado NÃO é versionado (contém dados reais; o repositório é
# público) — aplique-o no SQL Editor do Supabase após o sql/041 e após
# editar a linha EMAIL_RESPONSAVEL no topo do arquivo gerado.
#
# O que o script faz, por linha da planilha:
#  - cria o lic_processo (número LIC- gerado pelo trigger; criado_em =
#    data do evento mais antigo encontrado);
#  - quebra o diário da coluna "Situação" em movimentações 'andamento'
#    individuais, inferindo o ano das datas "dd/mm" pela data de
#    atualização (col. J) e mantendo a ordem cronológica;
#  - o trecho inicial sem data (resumo do estado atual) vira o último
#    andamento, datado pela "Data de Atualização";
#  - a coluna "Data do pregão e reagendamentos" vira uma movimentação
#    'agendamento_pregao' (data extraída quando possível);
#  - locais com grafias variantes são normalizados; fases da planilha
#    são mapeadas para o catálogo do sql/041; o que faltar é criado.
#  - Idempotente: pula processo já migrado (mesmo objeto + categoria).
# =====================================================================

import re
import sys
import unicodedata
from datetime import date, datetime

try:
    import openpyxl
except ImportError:
    sys.exit('Instale a dependência: pip install openpyxl')

ABAS = {
    'COMPRAS - ALMOX E SEDE': 'almox_sede',
    'COMPRAS - NUTRIÇÃO': 'nutricao',
    'COMPRAS - SERVIÇOS E OBRAS': 'servicos_obras',
}

PRIORIDADES = {'1': '1_secretario', '2': '2_alta', '3': '3_normal'}

# Fase da planilha (sem o prefixo "NN - ", em MAIÚSCULAS e sem acento)
# -> nome canônico do catálogo (seed do sql/041).
FASES = {
    'DFD': 'DFD',
    'TERMO DE REFERENCIA': 'Termo de referência',
    'COTACAO': 'Cotação',
    'REQUISICAO': 'Requisição',
    'ELABORAR EDITAL': 'Elaborar edital',
    'ENVIO AO JURIDICO': 'Envio ao jurídico',
    'AGENDAMENTO': 'Agendamento',
    'PREGAO AGENDADO': 'Pregão agendado',
    'PREGAO EM ANDAMENTO': 'Pregão em andamento',
    'HOMOLOGADO': 'Homologado',
    'ASSINATURA DE CONTRATO/ATA': 'Assinatura de contrato/ata',
    'LIBERADO PARA EMPENHO': 'Liberado para empenho',
    'AGUARDANDO OS': 'Aguardando OS',
    'EM EXECUCAO': 'Em execução',
    'CONTRATO FINALIZADO': 'Contrato finalizado',
    'SUSPENSO': 'Suspenso',
    'COMPRA DIRETA': 'Compra Direta',
}

# Grafias variantes de local -> sigla canônica (seed do sql/041).
LOCAIS = {
    'EDUC - ATAS E CONTRATOS': 'EDUC-ATAS E CONTRATOS',
    'EDUC-SUBLICITACOES': 'EDUC-SUB.LICITAÇÕES',
    'EDUC-SUB.LICITAÇÕES': 'EDUC-SUB.LICITAÇÕES',
    'EDUC-G.OBRAS': 'EDUC-OBRAS',
    'EDUC - ED. AMBIENTAL': 'EDUC-ED.AMBIENTAL',
    'EDUC-NUTRI. ESCOLAR': 'EDUC-NUTRI.ESCOLAR',
}

RE_ENTRADA = re.compile(r'(?=(?:^|\s)\d{1,2}/\d{1,2}(?:/\d{2,4})?\s*-\s)')
RE_DATA = re.compile(r'(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?')


def sem_acento(txt):
    return ''.join(c for c in unicodedata.normalize('NFD', txt)
                   if unicodedata.category(c) != 'Mn')


def sql_txt(v):
    if v is None or str(v).strip() == '':
        return 'null'
    return "'" + str(v).strip().replace("'", "''") + "'"


def limpar(v):
    """Texto de célula: colapsa os espaços em série usados como 'quebra'."""
    if v is None:
        return None
    t = re.sub(r'[ \t]{2,}', '\n', str(v)).strip()
    return t or None


def mapear_fase(bruto):
    if not bruto:
        return None
    chave = sem_acento(re.sub(r'^\s*\d+\s*-\s*', '', str(bruto)).strip()).upper()
    return FASES.get(chave)


def mapear_local(bruto):
    if not bruto or not str(bruto).strip():
        return None
    sigla = str(bruto).strip()
    return LOCAIS.get(sigla, sigla)


def mapear_prioridade(bruto):
    if not bruto:
        return '3_normal'
    return PRIORIDADES.get(str(bruto).strip()[:1], '3_normal')


def como_data(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def inferir_data(dd, mm, aa, referencia, anterior):
    """Datas 'dd/mm' do diário: ano inferido pela data de atualização,
    corrigido para manter a ordem cronológica das entradas."""
    if aa:
        ano = int(aa) + (2000 if int(aa) < 100 else 0)
        try:
            return date(ano, mm, dd)
        except ValueError:
            return None
    ano = referencia.year
    try:
        d = date(ano, mm, dd)
    except ValueError:
        return None
    if d > referencia:
        d = d.replace(year=ano - 1)
    while anterior and d < anterior:
        try:
            d = d.replace(year=d.year + 1)
        except ValueError:
            return None
        if d > referencia:            # inconsistência na planilha: mantém
            return min(d, referencia)
    return d


def quebrar_diario(texto, referencia):
    """Divide o diário em [(data|None, texto)] na ordem cronológica.
    O trecho inicial sem data (resumo atual) volta por último, datado
    pela data de atualização."""
    partes = [p.strip() for p in RE_ENTRADA.split(texto) if p and p.strip()]
    resumo = None
    entradas = []
    anterior = None
    for p in partes:
        m = RE_DATA.match(p)
        if not m:
            resumo = p if resumo is None else f'{resumo}\n{p}'
            continue
        d = inferir_data(int(m.group(1)), int(m.group(2)), m.group(3),
                         referencia, anterior)
        corpo = p[m.end():].lstrip(' -–').strip()
        if d is None:
            resumo = p if resumo is None else f'{resumo}\n{p}'
            continue
        anterior = d
        entradas.append((d, corpo or p))
    if resumo:
        entradas.append((referencia, f'[Resumo na migração] {resumo}'))
    return entradas


def processar(caminho):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    processos = []
    for aba, categoria in ABAS.items():
        if aba not in wb.sheetnames:
            print(f'AVISO: aba "{aba}" não encontrada — pulada.')
            continue
        for linha in wb[aba].iter_rows(min_row=2):
            v = [c.value for c in linha[:15]]
            objeto = limpar(v[0])
            if not objeto:
                continue
            fase = mapear_fase(v[6])
            if not fase:
                print(f'AVISO: fase "{v[6]}" sem mapa ({objeto[:40]}) — usando DFD.')
                fase = 'DFD'
            atualizacao = como_data(v[9]) or date.today()
            diario = quebrar_diario(limpar(v[10]) or '', atualizacao)
            pregao_txt = limpar(v[8])
            pregao_data = None
            if pregao_txt:
                m = RE_DATA.search(pregao_txt)
                if m and m.group(3):
                    pregao_data = inferir_data(int(m.group(1)), int(m.group(2)),
                                               m.group(3), atualizacao, None)
            processos.append({
                'categoria': categoria, 'objeto': objeto,
                'prioridade': mapear_prioridade(v[1]),
                'numero_requisicao': limpar(v[2]),
                'numero_processo_compra': limpar(v[3]),
                'numero_pregao': limpar(v[4]),
                'numero_processo_solar': limpar(v[5]),
                'fase': fase, 'local': mapear_local(v[7]),
                'pregao_txt': pregao_txt, 'pregao_data': pregao_data,
                'valor_requisicao': v[11], 'valor_arrematado': v[12],
                'valor_frustrado_deserto': v[14],
                'diario': diario, 'atualizacao': atualizacao,
            })
    return processos


def gerar_sql(processos):
    out = ["""-- SQL GERADO por scripts/migrar_licitacoes.py — NÃO VERSIONAR.
-- Aplicar no SQL Editor do Supabase DEPOIS do sql/041_licitacoes.sql.
-- 1) Edite EMAIL_RESPONSAVEL abaixo (autor da migração, usuário ativo).
-- 2) Rode o arquivo inteiro; reexecutar não duplica (pula existentes).
do $$
declare
  v_autor uuid;
  v_processo uuid;
  v_fase uuid;
  v_local uuid;
begin
  select id into v_autor from gestao.usuarios
   where email = 'EMAIL_RESPONSAVEL' and ativo;
  if v_autor is null then
    raise exception 'Edite EMAIL_RESPONSAVEL no topo deste arquivo.';
  end if;
"""]
    for p in processos:
        obj = sql_txt(p['objeto'])
        cat = sql_txt(p['categoria'])
        inicio = min([d for d, _ in p['diario']] or [p['atualizacao']])
        rotulo = ' '.join(p['objeto'].split())[:70]
        # A chave de idempotência inclui os números externos: a planilha
        # repete objeto na mesma categoria em edições de anos diferentes
        # (ex.: "Materiais de Limpeza" 2025 e 2026) — são processos
        # distintos e ambos devem entrar.
        out.append(f"""
  -- {rotulo}
  if not exists (select 1 from gestao.lic_processos
                  where objeto = {obj} and categoria = {cat}
                    and coalesce(numero_processo_compra, '') = coalesce({sql_txt(p['numero_processo_compra'])}, '')
                    and coalesce(numero_requisicao, '') = coalesce({sql_txt(p['numero_requisicao'])}, '')) then
    select id into v_fase from gestao.lic_fases where nome = {sql_txt(p['fase'])};
    if v_fase is null then
      raise exception 'Fase ausente do catálogo: %', {sql_txt(p['fase'])};
    end if;""")
        if p['local']:
            out.append(f"""    insert into gestao.lic_locais(sigla)
      select {sql_txt(p['local'])}
       where not exists (select 1 from gestao.lic_locais where sigla = {sql_txt(p['local'])});
    select id into v_local from gestao.lic_locais where sigla = {sql_txt(p['local'])};""")
        else:
            out.append('    v_local := null;')
        out.append(f"""    insert into gestao.lic_processos(
      objeto, categoria, prioridade, numero_requisicao, numero_processo_compra,
      numero_pregao, numero_processo_solar, fase_id, local_id, data_pregao,
      valor_requisicao, valor_arrematado, valor_frustrado_deserto,
      unidade_solicitante_id, criado_por, criado_em)
    values ({obj}, {cat}, {sql_txt(p['prioridade'])},
      {sql_txt(p['numero_requisicao'])}, {sql_txt(p['numero_processo_compra'])},
      {sql_txt(p['numero_pregao'])}, {sql_txt(p['numero_processo_solar'])},
      v_fase, v_local, {sql_txt(p['pregao_data'])},
      {p['valor_requisicao'] if p['valor_requisicao'] is not None else 'null'},
      {p['valor_arrematado'] if p['valor_arrematado'] is not None else 'null'},
      {p['valor_frustrado_deserto'] if p['valor_frustrado_deserto'] is not None else 'null'},
      (select unidade_id from gestao.usuarios where id = v_autor),
      v_autor, {sql_txt(inicio.isoformat())}::timestamptz)
    returning id into v_processo;
    insert into gestao.lic_movimentacoes(processo_id, autor_id, tipo, texto, criado_em)
    values (v_processo, v_autor, 'criacao',
      'Processo migrado da planilha de acompanhamento.',
      {sql_txt(inicio.isoformat())}::timestamptz);""")
        seq = 0
        for d, txt in p['diario']:
            seq += 1
            # minutos sequenciais preservam a ordem dentro do mesmo dia
            ts = f'{d.isoformat()}T12:{seq // 60:02d}:{seq % 60:02d}'
            out.append(f"""    insert into gestao.lic_movimentacoes(processo_id, autor_id, tipo, texto, criado_em)
    values (v_processo, v_autor, 'andamento', {sql_txt(txt)}, {sql_txt(ts)}::timestamptz);""")
        if p['pregao_txt']:
            out.append(f"""    insert into gestao.lic_movimentacoes(
      processo_id, autor_id, tipo, texto, data_pregao, criado_em)
    values (v_processo, v_autor, 'agendamento_pregao', {sql_txt(p['pregao_txt'])},
      {sql_txt(p['pregao_data'].isoformat()) if p['pregao_data'] else 'null'},
      {sql_txt(p['atualizacao'].isoformat())}::timestamptz);""")
        out.append('  end if;')
    out.append("end $$;\n")
    return '\n'.join(out)


def principal():
    if len(sys.argv) < 2:
        sys.exit('Uso: python3 scripts/migrar_licitacoes.py <planilha.xlsx> [saida.sql]')
    caminho = sys.argv[1]
    saida = sys.argv[2] if len(sys.argv) > 2 else 'migracao_licitacoes_dados.sql'
    processos = processar(caminho)
    with open(saida, 'w', encoding='utf-8') as f:
        f.write(gerar_sql(processos))
    total_movs = sum(len(p['diario']) + 1 + (1 if p['pregao_txt'] else 0)
                     for p in processos)
    print(f'{len(processos)} processos e {total_movs} movimentações -> {saida}')
    print('Edite EMAIL_RESPONSAVEL no arquivo gerado antes de aplicar.')


if __name__ == '__main__':
    principal()
