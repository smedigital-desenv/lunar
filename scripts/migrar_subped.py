#!/usr/bin/env python3
# =====================================================================
# migrar_subped.py — converte o "Painel de Tarefas" da Subsecretaria
# Pedagógica (planilha .xlsx) em SQL para o schema gestao.
#
# Desenho acertado com o usuário (2026-08-21):
#   · a DEMANDA fica na Subsecretaria Pedagógica (unidade + titular);
#   · o trabalho vai numa TAREFA para o "Responsável 2" da linha, que
#     pode ser uma gerência (usa-se o titular dela) ou uma pessoa;
#   · autor/solicitante = usuário de serviço "Dados importados", o mesmo
#     da migração de Licitações — ninguém real assina o que não escreveu;
#   · objeto_queixa (NOT NULL, sem coluna de origem) repete a Descrição;
#   · Link vira demandas.link_documentacao quando é URL; quando é só o
#     nome de um documento, fica preservado na descrição (o CHECK do
#     banco exige http(s)://);
#   · a coluna PUBLICAR é descartada.
#
# A tarefa é criada no formato "subtarefa direto na demanda" (sql/023):
# parent_id nulo e SEM mover responsavel_atual_id — que é o estado que a
# tela produz. fn_encaminhar moveria o responsável e tiraria a demanda da
# Subsecretaria, o oposto do combinado.
#
# O SQL gerado NÃO é versionado (dados reais, repositório público).
#
# Uso:
#   python3 scripts/migrar_subped.py <planilha.xlsx> [saida.sql]
# =====================================================================

import sys
import unicodedata
from datetime import datetime

import openpyxl

ABA = 'Painel de Tarefas'

# Colunas da planilha (índice 0)
C_DESCRICAO, C_PRIORIDADE, C_RESP1, C_RESP2 = 0, 1, 2, 3
C_STATUS, C_INICIO, C_FIM, C_TIPO = 4, 5, 6, 7
C_LINK, C_OBS, C_PRAZO, C_PUBLICAR = 8, 9, 10, 11

# Status da planilha -> situacao da demanda.
SITUACAO = {
    'em execução': 'em_andamento',
    'não iniciada': 'aberta',
    'concluída': 'concluida',
}

# Unidade que recebe todas as demandas. Resolvida por NOME no banco: o
# organograma oficial (sql/038) não é versionado, então o script não pode
# embutir UUID nenhum.
UNIDADE_TITULAR = 'SUBSECRETARIA PEDAGÓGICA'

# Usuário de serviço da migração (criado na recarga de Licitações).
EMAIL_AUTOR = 'dados.importados@educacao.pmrp.sp.gov.br'

# Grafias da planilha que não batem com o organograma oficial. A busca já
# ignora acento e caixa; isto é para diferença de PALAVRAS mesmo.
#
# Preencher a partir da consulta 2 de scripts/diagnostico_subped.sql, que
# lista os nomes reais sob a Subsecretaria. Deixar em branco não é opção
# silenciosa: o nome que não resolver aborta a migração inteira.
MAPA_UNIDADES = {
    # Confirmado como divergente em 2026-08-21 (a migração abortou nele).
    # O nome REAL sai da consulta 2 do diagnóstico — não preencher por
    # palpite: um nome errado que exista manda a tarefa para a unidade
    # errada, e aí a carga passa em silêncio.
    # 'GERÊNCIA DO CENTRO DE FORMAÇÃO PAULO FREIRE': '<nome exato no organograma>',
}

# Mesmo mecanismo para pessoas (casamento por nome completo).
MAPA_PESSOAS = {
    # 'NOME NA PLANILHA': 'Nome exato no cadastro',
}


def norm(expr):
    """Expressão SQL que normaliza um nome para comparação: sem acento,
    sem caixa, sem aspas decorativas e com espaços colapsados.

    O organograma escreve Gerência do Centro de Formação "Paulo Freire"
    entre aspas e a planilha não — foi o que abortou a carga de
    2026-08-21. Aspas são enfeite de grafia, não identidade da unidade;
    resolver aqui evita um mapa por unidade aspeada. Só aspas: tirar
    hífen ou ponto criaria equivalências que ninguém pediu.
    """
    sem_aspas = f"""translate({expr}, '"''´`', '')"""
    return f"btrim(regexp_replace(upper(unaccent({sem_aspas})), '\\s+', ' ', 'g'))"


def mapear(nome, mapa):
    """Aplica o mapa de grafias, comparando sem acento e sem caixa."""
    if not nome:
        return nome
    alvo = sem_acento(nome).upper().strip()
    for de, para in mapa.items():
        if sem_acento(de).upper().strip() == alvo:
            return para
    return nome


def sem_acento(txt):
    return ''.join(c for c in unicodedata.normalize('NFD', str(txt))
                   if unicodedata.category(c) != 'Mn')


def sql_txt(v):
    """Literal SQL. None vira NULL; aspas simples são dobradas."""
    if v is None:
        return 'null'
    return "'" + str(v).replace("'", "''") + "'"


def limpar(v):
    if v is None:
        return None
    t = ' '.join(str(v).split())
    return t or None


def como_data(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    try:
        return datetime.fromisoformat(str(v)).date()
    except ValueError:
        return None


def e_unidade(nome):
    """A coluna "Responsável 2" mistura gerências e pessoas. Distinguir
    pelo prefixo é o que a planilha permite — nomes de unidade sempre
    começam por GERÊNCIA/SUBSECRETARIA/SEÇÃO/NÚCLEO."""
    if not nome:
        return False
    n = sem_acento(nome).upper()
    return n.startswith(('GERENCIA', 'SUBSECRETARIA', 'SECAO', 'NUCLEO', 'GABINETE'))


def montar_descricao(obs, link, link_e_url):
    """Descrição = Observações + o nome do documento quando ele não é uma
    URL (esse não cabe em link_documentacao por causa do CHECK)."""
    partes = []
    if obs:
        partes.append(obs)
    if link and not link_e_url:
        partes.append(f'Documento de referência: {link}')
    return '\n\n'.join(partes) or None


def processar(caminho):
    wb = openpyxl.load_workbook(caminho, data_only=True)
    if ABA not in wb.sheetnames:
        sys.exit(f'Aba {ABA!r} não encontrada. Abas: {wb.sheetnames}')
    ws = wb[ABA]

    itens = []
    for linha, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if linha == 1:
            continue                       # cabeçalho
        titulo = limpar(row[C_DESCRICAO])
        if not titulo:
            continue                       # linha vazia

        status = (limpar(row[C_STATUS]) or '').lower()
        situacao = SITUACAO.get(status)
        if situacao is None:
            sys.exit(f'Linha {linha}: status desconhecido {status!r}. '
                     f'Conhecidos: {sorted(SITUACAO)}')

        obs = limpar(row[C_OBS])
        link = limpar(row[C_LINK])
        link_e_url = bool(link and link.lower().startswith(('http://', 'https://')))
        fim = como_data(row[C_FIM])

        if situacao == 'concluida' and not obs:
            # chk_demanda_conclusao: concluída exige conclusão preenchida.
            sys.exit(f'Linha {linha}: {titulo!r} está Concluída sem Observações — '
                     'não há texto para a conclusão exigida pelo banco.')

        itens.append({
            'linha': linha,
            'titulo': titulo,
            'situacao': situacao,
            'descricao': montar_descricao(obs, link, link_e_url),
            # Concluída: a observação vira a conclusão e o término, a data.
            # Em aberto: o término é o prazo.
            'conclusao': obs if situacao == 'concluida' else None,
            'data_conclusao': fim if situacao == 'concluida' else None,
            'prazo': fim if situacao != 'concluida' else None,
            'inicio': como_data(row[C_INICIO]),
            'link': link if link_e_url else None,
            # `responsavel2` é o texto original (vai para a descrição da
            # tarefa); `resp2_banco` é o nome já mapeado, usado na busca.
            'responsavel2': limpar(row[C_RESP2]),
            'resp2_banco': mapear(
                limpar(row[C_RESP2]),
                MAPA_UNIDADES if e_unidade(limpar(row[C_RESP2])) else MAPA_PESSOAS),
        })
    return itens


def gerar_sql(itens, planilha):
    hoje = datetime.now().date().isoformat()
    out = [f"""-- Migração do Painel de Tarefas da Subsecretaria Pedagógica.
-- Gerado por scripts/migrar_subped.py a partir de {planilha}
-- {len(itens)} demandas. NÃO versionar: contém dados reais.
--
-- Pré-requisito: rodar sql/044_demandas_link_documentacao.sql antes.
-- Idempotente por (titulo, unidade_responsavel_id): rodar de novo não
-- duplica. Nada é apagado (regra 1).

set search_path = gestao, extensions, public;

do $$
declare
  v_autor    uuid;
  v_unidade  uuid;
  v_titular  uuid;
  v_resp     uuid;
  v_uresp    uuid;
  v_demanda  uuid;
  v_tarefa   uuid;
  v_criado   timestamptz;
  v_desc     text;
  v_sobrou   int := 0;   -- pessoas não localizadas que caíram na Subsecretaria
begin
  select id into v_autor from gestao.usuarios where email = {sql_txt(EMAIL_AUTOR)};
  if v_autor is null then
    raise exception 'Usuário de serviço da migração não existe: %', {sql_txt(EMAIL_AUTOR)};
  end if;

  select id into v_unidade from gestao.unidades_organizacionais
   where {norm('nome')} = {norm(sql_txt(UNIDADE_TITULAR))} and ativo;
  if v_unidade is null then
    raise exception 'Unidade não encontrada no organograma: %', {sql_txt(UNIDADE_TITULAR)};
  end if;

  -- Titular da Subsecretaria: quem está lotado nela com o maior nível.
  select u.id into v_titular
    from gestao.usuarios u join gestao.perfis p on p.codigo = u.perfil
   where u.unidade_id = v_unidade and u.ativo
   order by p.nivel desc, u.criado_em limit 1;
  if v_titular is null then
    raise exception 'Nenhum usuário ativo lotado em %', {sql_txt(UNIDADE_TITULAR)};
  end if;
"""]

    for it in itens:
        rotulo = it['titulo'][:70]
        out.append(f"""
  -- linha {it['linha']}: {rotulo}
  if not exists (select 1 from gestao.demandas
                  where titulo = {sql_txt(it['titulo'])}
                    and unidade_responsavel_id = v_unidade) then
    v_criado := coalesce({sql_txt(it['inicio'].isoformat()) if it['inicio'] else 'null'}::timestamptz,
                         now());
    insert into gestao.demandas(
      titulo, objeto_queixa, descricao, situacao, prioridade, prazo,
      conclusao, data_conclusao, link_documentacao, sigilo,
      solicitante_id, responsavel_atual_id, unidade_responsavel_id,
      criado_por, criado_em)
    values (
      {sql_txt(it['titulo'])}, {sql_txt(it['titulo'])}, {sql_txt(it['descricao'])},
      {sql_txt(it['situacao'])}, 'normal',
      {sql_txt(it['prazo'].isoformat()) if it['prazo'] else 'null'}::date,
      {sql_txt(it['conclusao'])},
      {sql_txt(it['data_conclusao'].isoformat()) if it['data_conclusao'] else 'null'}::timestamptz,
      {sql_txt(it['link'])}, 'normal',
      v_autor, v_titular, v_unidade, v_autor, v_criado)
    returning id into v_demanda;

    -- Mesmos participantes que fn_criar_demanda registraria.
    perform gestao.fn_garantir_participante(v_demanda, v_autor, 'solicitante', v_autor);
    perform gestao.fn_garantir_participante(v_demanda, v_titular, 'responsavel', v_autor);

    insert into gestao.movimentacoes(
      demanda_id, autor_id, tipo, texto, situacao_nova, destinatario_id,
      prioridade, criado_em)
    values (v_demanda, v_autor, 'criacao',
      'Demanda migrada do Painel de Tarefas da Subsecretaria Pedagógica.',
      {sql_txt(it['situacao'])}, v_titular, 'normal', v_criado);""")

        r2 = it['responsavel2']          # texto original, para a descrição
        r2b = it['resp2_banco']          # nome já mapeado, para a busca
        if r2 and e_unidade(r2):
            out.append(f"""
    -- Responsável 2 é unidade: a tarefa vai para o titular dela.
    select id into v_uresp from gestao.unidades_organizacionais
     where {norm('nome')} = {norm(sql_txt(r2b))} and ativo;
    if v_uresp is null then
      raise exception 'Linha {it['linha']}: unidade não encontrada: % — mapeie a grafia em MAPA_UNIDADES (scripts/migrar_subped.py).', {sql_txt(r2b)};
    end if;
    select u.id into v_resp
      from gestao.usuarios u join gestao.perfis p on p.codigo = u.perfil
     where u.unidade_id = v_uresp and u.ativo
     order by p.nivel desc, u.criado_em limit 1;
    if v_resp is null then
      raise exception 'Linha {it['linha']}: sem titular ativo em %', {sql_txt(r2b)};
    end if;
    v_desc := {sql_txt(f'Responsável na planilha de origem: {r2}.')};""")
        elif r2:
            out.append(f"""
    -- Responsável 2 é pessoa. Só os 52 titulares do organograma foram
    -- cadastrados (sql/039), então quem não chefia unidade pode não ter
    -- conta. Decisão do usuário (2026-08-21): não abortar — a tarefa vai
    -- para a titular da Subsecretaria, com o nome de origem registrado,
    -- e ela redistribui quando a conta existir. Inativo conta como
    -- ausente: atribuir a quem saiu ressuscitaria trabalho.
    select id, unidade_id into v_resp, v_uresp from gestao.usuarios
     where {norm('nome')} = {norm(sql_txt(r2b))} and ativo;
    v_desc := {sql_txt(f'Responsável na planilha de origem: {r2}.')};
    if v_resp is null then
      v_resp   := v_titular;
      v_uresp  := v_unidade;
      v_sobrou := v_sobrou + 1;
      v_desc   := v_desc || ' Conta não localizada (ou inativa) no cadastro;'
                         || ' atribuída à Subsecretaria para redistribuição.';
      raise notice 'Linha {it['linha']}: % sem conta ativa — tarefa para a Subsecretaria.', {sql_txt(r2b)};
    end if;""")
        else:
            out.append("""
    v_resp := null;   -- linha sem Responsável 2: fica só a demanda
    v_desc := null;""")

        # Demanda concluída não pode deixar tarefa aberta: ela ficaria para
        # sempre na caixa de entrada da gerência e apareceria como marco em
        # aberto no cronograma de "Meus processos". chk_tarefa_conclusao
        # exige o texto da conclusão junto.
        concluida = it['situacao'] == 'concluida'
        sit_tarefa = 'concluida' if concluida else 'aberta'
        out.append(f"""
    if v_resp is not null then
      insert into gestao.tarefas(
        demanda_id, parent_id, titulo, descricao, responsavel_id,
        unidade_responsavel_id, situacao, prioridade, prazo,
        conclusao, data_conclusao, criado_por, criado_em)
      values (v_demanda, null, {sql_txt(it['titulo'])}, v_desc,
        v_resp, v_uresp, {sql_txt(sit_tarefa)}, 'normal',
        {sql_txt(it['prazo'].isoformat()) if it['prazo'] else 'null'}::date,
        {sql_txt(it['conclusao'])},
        {sql_txt(it['data_conclusao'].isoformat()) if it['data_conclusao'] else 'null'}::timestamptz,
        v_autor, v_criado)
      returning id into v_tarefa;

      insert into gestao.movimentacoes(
        demanda_id, tarefa_id, autor_id, tipo, texto, situacao_nova,
        destinatario_id, prioridade, criado_em)
      values (v_demanda, v_tarefa, v_autor, 'subtarefa', {sql_txt(it['titulo'])},
        {sql_txt(sit_tarefa)}, v_resp, 'normal', v_criado);
    end if;
  end if;""")

    out.append(f"""
  raise notice 'Migração concluída em {hoje}. % tarefa(s) sem conta própria foram para a Subsecretaria.', v_sobrou;
end $$;

-- Conferência (rode depois):
--   select count(*) from gestao.demandas d
--     join gestao.unidades_organizacionais u on u.id = d.unidade_responsavel_id
--    where {norm('u.nome')} = {norm(sql_txt(UNIDADE_TITULAR))};
--   -- esperado: {len(itens)}
""")
    return '\n'.join(out)


def gerar_conferencia(itens):
    """SQL só de leitura, para rodar ANTES: mostra quais nomes da planilha
    o banco encontra. Sem isso a migração só falha no primeiro nome ruim,
    um de cada vez."""
    nomes_un = sorted({it['resp2_banco'] for it in itens
                       if it['responsavel2'] and e_unidade(it['responsavel2'])})
    nomes_pe = sorted({it['resp2_banco'] for it in itens
                       if it['responsavel2'] and not e_unidade(it['responsavel2'])})
    un = ', '.join(f'({sql_txt(n)})' for n in [UNIDADE_TITULAR] + nomes_un)
    pe = ', '.join(f'({sql_txt(n)})' for n in nomes_pe)
    return f"""-- Conferência de nomes — SÓ LEITURA. Rode ANTES da migração.
-- Toda linha com achou = false precisa de correção (na planilha ou no
-- cadastro) antes de migrar.
set search_path = gestao, extensions, public;

select 'unidade' as tipo, v.nome,
       (u.id is not null) as achou,
       (select count(*) from gestao.usuarios x
         where x.unidade_id = u.id and x.ativo) as pessoas_ativas
  from (values {un}) as v(nome)
  left join gestao.unidades_organizacionais u
    on {norm('u.nome')} = {norm('v.nome')} and u.ativo
union all
select 'pessoa', v.nome, (p.id is not null), null
  from (values {pe}) as v(nome)
  left join gestao.usuarios p
    on {norm('p.nome')} = {norm('v.nome')} and p.ativo
order by achou, tipo, nome;

-- Usuário de serviço da migração:
select 'autor' as tipo, {sql_txt(EMAIL_AUTOR)} as nome,
       exists (select 1 from gestao.usuarios where email = {sql_txt(EMAIL_AUTOR)}) as achou;
"""


def principal():
    if len(sys.argv) < 2:
        sys.exit('Uso: python3 scripts/migrar_subped.py <planilha.xlsx> [saida.sql]')
    planilha = sys.argv[1]
    saida = sys.argv[2] if len(sys.argv) > 2 else 'migracao_subped_dados.sql'

    itens = processar(planilha)
    with open(saida, 'w', encoding='utf-8') as f:
        f.write(gerar_sql(itens, planilha))
    conf = saida.replace('.sql', '_conferencia.sql')
    with open(conf, 'w', encoding='utf-8') as f:
        f.write(gerar_conferencia(itens))

    com_tarefa = sum(1 for i in itens if i['responsavel2'])
    print(f'{len(itens)} demandas ({com_tarefa} com tarefa) -> {saida}')
    print(f'conferência de nomes -> {conf}')
    for s in sorted({i['situacao'] for i in itens}):
        print(f"  {s}: {sum(1 for i in itens if i['situacao'] == s)}")
    print(f"  com link_documentacao: {sum(1 for i in itens if i['link'])}")
    sem_r2 = [i['titulo'] for i in itens if not i['responsavel2']]
    if sem_r2:
        print(f'  SEM tarefa (sem Responsável 2): {sem_r2}')


if __name__ == '__main__':
    principal()
